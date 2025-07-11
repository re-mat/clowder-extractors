#!/usr/bin/env python
import json
import sys
from typing import Tuple, List, Dict

# from .chemistry import Monomer, ChemDB, Catalyst, Inhibitor, Solvent, Additive, Initiator
from clowder_extractors.experiment_from_excel.chemistry import (
    Monomer,
    ChemDB,
    Catalyst,
    Inhibitor,
    Solvent,
    Additive,
    Initiator,
)
import logging

import pyclowder.files
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pyclowder.extractors import Extractor
from pyclowder.utils import CheckMessage

moles_format = "{:.2e}"


def microliters_to_milli(value):
    if value and value != "-":
        return float(value) / 1000.0
    else:
        return value


def mass_volume_conversion(value):
    if value == "-":
        return None
    return value


def find_mass_column(row):
    if "Measured mass (g)" in row:
        return "Measured mass (g)"
    elif "Measured mass (mg)" in row:
        return "Measured mass (mg)"
    else:
        return None


def find_volume_column(row):
    if "Measured volume (μL)" in row:
        return "Measured volume (μL)"


def is_row_empty(row, input_title: str):

    if "SMILES" not in row:
        return True

    mass_key = find_mass_column(row)
    volume_key = find_volume_column(row)

    smiles = row["SMILES"] if row["SMILES"] else None
    name = row["Name"] if row["Name"] else None
    mass = mass_volume_conversion(row[mass_key]) if mass_key else None
    volume = mass_volume_conversion(row[volume_key]) if volume_key else None

    if name and smiles:
        if mass and volume:
            raise ValueError(
                f"On {input_title} tab: Only specify one of mass or volume in "
                + str(row)
            )
        elif not mass and not volume:
            raise ValueError(
                f"On {input_title} tab: Volume or mass must be specified in " + str(row)
            )

    if mass_key and volume_key:
        if not name and not smiles and not mass and not volume:
            return True
    elif mass_key:
        if not name and not smiles and not mass:
            return True
    elif volume_key:
        if not name and not smiles and not volume:
            return True

    if not all([row["Name"], row["SMILES"]]):
        raise ValueError(f"On {input_title} tab: Missing Name or SMILES in " + str(row))

    return False


def compute_values(inputs: dict, inputs_procedure: dict):
    # First, create lists of input-specific chemistry converters with the observed values
    db = ChemDB()

    # MONOMERS
    db.exists([compound["SMILES"] for compound in inputs["monomers"]])
    monomers = {}
    for compound in inputs["monomers"]:
        if "Measured mass (g)" in compound:
            measured_mass = compound["Measured mass (g)"]
        elif "Measured mass (mg)" in compound:
            # For monomer alone we will use (g)
            measured_mass = compound["Measured mass (mg)"] / 1000.0  # Convert mg to g
            compound["Measured mass (g)"] = (
                measured_mass  # Add the key to the compound dict for further use
            )
        else:
            measured_mass = None

        monomers[compound["SMILES"]] = Monomer(
            compound["SMILES"],
            db,
            mass_volume_conversion(measured_mass),
            microliters_to_milli(
                mass_volume_conversion(compound["Measured volume (μL)"])
            ),
        )

    # CATALYSTS
    db.exists([compound["SMILES"] for compound in inputs["catalysts"]])
    catalysts = {}
    for compound in inputs["catalysts"]:
        if "Measured mass (g)" in compound:
            measured_mass_g = compound["Measured mass (g)"]
            compound["Measured mass (mg)"] = measured_mass_g * 1000.0  # Convert g to mg
        elif "Measured mass (mg)" in compound:
            measured_mass_g = compound["Measured mass (mg)"] / 1000.0
        else:
            measured_mass_g = None

        catalysts[compound["SMILES"]] = Catalyst(
            compound["SMILES"],
            db,
            mass_volume_conversion(measured_mass_g),
            None,
        )

    # INHIBITORS
    db.exists([compound["SMILES"] for compound in inputs["inhibitors"]])
    inhibitors = {}
    for compound in inputs["inhibitors"]:
        if "Measured mass (g)" in compound:
            measured_mass = compound["Measured mass (g)"]
        elif "Measured mass (mg)" in compound:
            measured_mass = compound["Measured mass (mg)"] / 1000.0
            compound["Measured mass (g)"] = (
                measured_mass  # Add the key to the compound dict for further use
            )
        else:
            measured_mass = None  # Handle missing mass case

        inhibitors[compound["SMILES"]] = Inhibitor(
            compound["SMILES"],
            db,
            mass_volume_conversion(measured_mass),
            microliters_to_milli(
                mass_volume_conversion(compound["Measured volume (μL)"])
            ),
        )

    # ADDITIVES
    # will use (g) for calculation
    db.exists([compound["SMILES"] for compound in inputs["additives"]])
    additives = {}
    for compound in inputs["additives"]:
        if "Measured mass (g)" in compound:
            measured_mass = compound["Measured mass (g)"]
        elif "Measured mass (mg)" in compound:
            measured_mass = compound["Measured mass (mg)"] / 1000.0  # Convert mg to g
            compound["Measured mass (g)"] = measured_mass
        else:
            measured_mass = None

        additives[compound["SMILES"]] = Additive(
            compound["SMILES"],
            db,
            mass_volume_conversion(measured_mass),
            microliters_to_milli(
                mass_volume_conversion(compound["Measured volume (μL)"])
            ),
        )

    # SOLVENTS
    # will use (g) for calculation
    db.exists([compound["SMILES"] for compound in inputs["solvents"]])
    solvents = {}
    for compound in inputs["solvents"]:
        if "Measured mass (g)" in compound:
            measured_mass = compound["Measured mass (g)"]
            compound["Measured mass (mg)"] = (
                measured_mass * 1000.0
            )  # to display in meta-data for measured mass

        elif compound.get("Measured mass (mg)", None):
            measured_mass = compound["Measured mass (mg)"] / 1000.0
            compound["Measured mass (g)"] = measured_mass
        else:
            measured_mass = None

        solvents[compound["SMILES"]] = Solvent(
            compound["SMILES"],
            db,
            mass_volume_conversion(measured_mass),
            microliters_to_milli(
                mass_volume_conversion(compound["Measured volume (μL)"])
            ),
        )

    # INITIATORS
    # will use  mg for calculation
    db.exists([compound["SMILES"] for compound in inputs["chemical initiation"]])
    initiators = {}
    for compound in inputs["chemical initiation"]:
        if "Measured mass (g)" in compound:
            measured_mass = compound["Measured mass (g)"]
        elif "Measured mass (mg)" in compound:
            measured_mass = compound["Measured mass (mg)"] / 1000.0  # Convert mg to g
            compound["Measured mass (g)"] = measured_mass
        else:
            measured_mass = None

        initiators[compound["SMILES"]] = Initiator(
            compound["SMILES"],
            db,
            mass=mass_volume_conversion(measured_mass),
            volume=microliters_to_milli(
                mass_volume_conversion(compound["Measured volume (μL)"])
            ),
        )

    total_initiator_catalyst_moles = sum(
        initiator.moles()
        for initiator in initiators.values()
        if initiator.role == Initiator.InitiatorRole.Catalyst
    )
    total_initiator_catalyst_mg = sum(
        initiator.mass
        for initiator in initiators.values()
        if initiator.role == Initiator.InitiatorRole.Catalyst
    )
    total_initiator_solvent_microliters = sum(
        initiator.volume
        for initiator in initiators.values()
        if initiator.role == Initiator.InitiatorRole.Solvent
    )

    # Now compute derived values (which requires knowledge of all of the inputs)
    monomer2 = [
        {
            "name": monomer["Name"],
            "SMILES": monomer["SMILES"],
            "Measured mass (g)": monomer["Measured mass (g)"],
            "Measured volume (μL)": monomer["Measured volume (μL)"],
            "Computed mass (g)": monomers[monomer["SMILES"]].mass,
            "Molecular Weight (g/mol)": monomers[monomer["SMILES"]].molecular_weight,
            "Moles": moles_format.format(monomers[monomer["SMILES"]].moles()),
            "Monomer mol%": monomers[monomer["SMILES"]].monomer_mol_percent(
                monomers.values()
            ),
        }
        for monomer in inputs["monomers"]
    ]

    catalyst2 = [
        {
            "name": catalyst["Name"],
            "SMILES": catalyst["SMILES"],
            "Measured mass (mg)": catalyst["Measured mass (mg)"],
            "Computed mass (g)": catalysts[catalyst["SMILES"]].mass,
            "Molecular Weight (g/mol)": catalysts[catalyst["SMILES"]].molecular_weight,
            "Moles": moles_format.format(catalysts[catalyst["SMILES"]].moles()),
            "Monomer:Catalyst molar ratio": catalysts[
                catalyst["SMILES"]
            ].catalyst_monomer_molar_ratio(monomers.values()),
        }
        for catalyst in inputs["catalysts"]
    ]

    inhibitor2 = [
        {
            "name": inhibitor["Name"],
            "SMILES": inhibitor["SMILES"],
            "Measured volume (μL)": inhibitor["Measured volume (μL)"],
            "Density": inhibitors[inhibitor["SMILES"]].density,
            "Computed mass (g)": inhibitors[inhibitor["SMILES"]].mass,
            "Molecular Weight (g/mol)": inhibitors[
                inhibitor["SMILES"]
            ].molecular_weight,
            "Moles": moles_format.format(inhibitors[inhibitor["SMILES"]].moles()),
            "Inhibitor:Catalyst molar ratio": inhibitors[
                inhibitor["SMILES"]
            ].inhibitor_catalyst_molar_ratio(list(catalysts.values())),
        }
        for inhibitor in inputs["inhibitors"]
    ]

    additive2 = [
        {
            "name": additive["Name"],
            "SMILES": additive["SMILES"],
            "Measured mass (g)": additive["Measured mass (g)"],
            "Measured volume (μL)": additive["Measured volume (μL)"],
            "Computed mass (g)": additives[additive["SMILES"]].mass,
            "Molecular Weight (g/mol)": additives[additive["SMILES"]].molecular_weight,
            "Moles": moles_format.format(additives[additive["SMILES"]].moles()),
            "Wt Percent of Additives": additives[
                additive["SMILES"]
            ].additive_weight_percent(
                list(additives.values()),
                list(monomers.values()),
                list(catalysts.values()),
                list(solvents.values()),
            ),
        }
        for additive in inputs["additives"]
    ]

    solvents2 = [
        {
            "name": solvent["Name"],
            "SMILES": solvent["SMILES"],
            "Measured mass (mg)": solvent["Measured mass (mg)"],
            "Measured volume (μL)": solvent["Measured volume (μL)"],
            "Computed mass (g)": solvents[solvent["SMILES"]].mass,
            "Molecular Weight (g/mol)": solvents[solvent["SMILES"]].molecular_weight,
            "Moles": moles_format.format(solvents[solvent["SMILES"]].moles()),
            "Solvent concentration (mL/g)": solvents[
                solvent["SMILES"]
            ].solvent_concentration(list(catalysts.values())[0]),
        }
        for solvent in inputs["solvents"]
    ]

    chemical_initiation2 = [
        {
            "name": chemical_initiation["Name"],
            "SMILES": chemical_initiation["SMILES"],
            "Role": initiators[chemical_initiation["SMILES"]].role.value,
            "Measured mass (mg)": chemical_initiation["Measured mass (mg)"],
            "Measured volume (μL)": chemical_initiation["Measured volume (μL)"],
            "Molecular Weight (g/mol)": initiators[
                chemical_initiation["SMILES"]
            ].molecular_weight,
            "Moles": moles_format.format(
                initiators[chemical_initiation["SMILES"]].moles()
            ),
        }
        for chemical_initiation in inputs["chemical initiation"]
    ]

    inputs["monomers"] = {
        "monomer-inputs": monomer2,
        "monomer-procedure": inputs_procedure["monomers"],
    }

    inputs["catalysts"] = {
        "catalyst-inputs": catalyst2,
        "catalyst-procedure": inputs_procedure["catalysts"],
    }

    inputs["inhibitors"] = {
        "inhibitor-inputs": inhibitor2,
        "inhibitor-procedure": inputs_procedure["inhibitors"],
    }

    inputs["additives"] = {
        "additive-inputs": additive2,
        "additive-procedure": inputs_procedure["additives"],
    }

    inputs["solvents"] = {
        "solvent-inputs": solvents2,
        "solvent-procedure": inputs_procedure["solvents"],
    }

    # Avoid divide by zero errors if the chemical initiation tab is not used
    if total_initiator_solvent_microliters:
        inputs["chemical initiation"] = {
            "initiator-catalyst-solvent-concentration-mg/microL": total_initiator_catalyst_mg
            / total_initiator_solvent_microliters,
            "initiator-catalyst-solvent-concentration-moles/L": total_initiator_catalyst_moles
            / (total_initiator_solvent_microliters * 10e6),
            "initiator-inputs": chemical_initiation2,
            "initiator-procedure": inputs_procedure["chemical initiation"],
        }


def read_inputs_from_worksheet(ws: Worksheet) -> Tuple[List[Dict], Dict]:
    # The inputs sheets contain rows of inputs and then a procedure block
    # that applies to all of the inputs of that type
    inputs = []
    procedure = {}
    inside_procedure_block = False
    headers = [col.value for col in list(ws.rows)[0]]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "PROCEDURE":
            inside_procedure_block = True
            continue

        if not inside_procedure_block:
            input_properties = {key: cell.value for key, cell in zip(headers, row)}

            if not is_row_empty(input_properties, input_title=ws.title):
                inputs.append(input_properties)
        else:
            if row[0].value:
                procedure[row[0].value] = row[1].value

    return inputs, procedure


def read_procedure_from_worksheet(ws: Worksheet) -> dict:
    procedure = {}
    for row in ws.iter_rows():
        procedure[row[0].value] = row[1].value

    return procedure


def read_batch_id(wb: Workbook) -> str:
    batch_id_range = wb.defined_names["BatchID"].value.split("!")
    batch_id_cell = wb[batch_id_range[0]][batch_id_range[1]]
    return batch_id_cell.value


def excel_to_json(path):
    logging.getLogger("__main__")

    wb = load_workbook(filename=path, data_only=True)
    # Find the data input spreadsheet version
    ss_version = [
        prop.value for prop in wb.custom_doc_props.props if prop.name == "File Version"
    ][0]

    if ss_version != "3.0":
        raise ValueError(
            f"This extractor is not compatible with spreadsheet version {ss_version}"
        )

    inputs = {}
    batch_id = read_batch_id(wb)
    procedure = {}
    fromp_measurements = {}
    inputs_procedure = {}
    initiation_tabs = {"THERMAL": "thermal initiation", "PHOTO": "photo initiation"}

    fromp_properties = [
        "Select Geometry from geometries tab",
        "Geometry - Select from library",
        "Resin height (mm)",
        "Diameter (mm)",
        "Thickness (mm)",
        "Tube length (mm)",
        "Empty-dim",
    ]
    # There are multiple sheets in this workbook. Some describe the inputs some
    # are just procedure. The Geometry sheet is just a library of geometries
    for sheet in wb.sheetnames:
        if sheet == "geometries":
            pass  # This sheet is just a library of geometries
        elif sheet in [
            "general",
            "thermal initiation",
            "photo initiation",
            "photo control",
        ]:
            procedure[sheet] = read_procedure_from_worksheet(wb[sheet])
        elif sheet == "FROMP Measurements":
            fromp_measurements = read_procedure_from_worksheet(wb[sheet])
        else:
            # The inputs sheets have both the list of inputs and procedure data
            inputs[sheet], inputs_procedure[sheet] = read_inputs_from_worksheet(
                wb[sheet]
            )

    compute_values(inputs, inputs_procedure)

    # Clean up irrelevant procedure data
    if procedure["general"]["Photocontrol?"] == "NO":
        del procedure["photo control"]

    # Remove entries for the initiation methods not used in this procedure
    # (Chemical initiation data is under the inputs section)
    unused_tabs = [
        value
        for key, value in initiation_tabs.items()
        if key != procedure["general"]["Initiation method"]
    ]
    for tab in unused_tabs:
        del procedure[tab]

    if procedure["general"]["Initiation method"] == "CHEMICAL":
        procedure["chemical initiation"] = inputs["chemical initiation"]
        del inputs["chemical initiation"]
    else:
        del inputs["chemical initiation"]

    # If not FROMP, then there is no initiation and we don't care about geometry
    if procedure["general"]["Type of polymerization"] == "NONE":
        for fromp_property in fromp_properties:
            procedure.pop(fromp_property, None)

    result = {
        "Batch ID": batch_id,
        "procedure": procedure,
        "inputs": inputs,
    }

    if procedure["general"]["Type of polymerization"] == "FROMP":
        result["FROMP Measurements"] = fromp_measurements

    return result


class ExperimentFromExcel(Extractor):
    def __init__(self):
        Extractor.__init__(self)
        # parse command line and load default logging configuration
        self.setup()
        # setup logging for the extractor
        logging.getLogger("pyclowder").setLevel(logging.DEBUG)
        logging.getLogger("__main__").setLevel(logging.DEBUG)

    def check_message(self, connector, host, secret_key, resource, parameters):
        logging.getLogger(__name__).debug("default check message : " + str(parameters))
        return CheckMessage.download

    def process_message(self, connector, host, secret_key, resource, parameters):
        logger = logging.getLogger("__main__")
        experiment = excel_to_json(resource["local_paths"][0])
        logger.debug(experiment)

        # store results as metadata
        metadata = {
            "@context": ["https://clowder.ncsa.illinois.edu/contexts/metadata.jsonld"],
            "dataset_id": resource["parent"].get("id", None),
            "content": experiment,
            "agent": {
                "@type": "cat:extractor",
                "extractor_id": host + "api/extractors/" + self.extractor_info["name"],
            },
        }

        print(
            json.loads(json.dumps(metadata, default=str, ensure_ascii=False, indent=4))
        )
        pyclowder.datasets.upload_metadata(
            connector,
            host,
            secret_key,
            resource["parent"].get("id", None),
            json.loads(json.dumps(metadata, default=str, ensure_ascii=False)),
        )


def main():
    if len(sys.argv) == 2:
        experiment = excel_to_json(sys.argv[1])
        print(json.dumps(experiment, indent=4, default=str, ensure_ascii=False))
    else:
        extractor = ExperimentFromExcel()
        extractor.start()


if __name__ == "__main__":
    main()

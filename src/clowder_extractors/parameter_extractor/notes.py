import logging
from datetime import datetime

from clowder_extractors.experiment_from_excel.chemistry import ChemDB
from openpyxl import load_workbook


logger = logging.getLogger(__name__)


class Notes:
    def __init__(self, parameters: dict):
        self.parameters = parameters
        self.notes = self.extract_notes_field()
        self.chemDB = ChemDB()
        self.path = ""

    def extract_notes_field(self) -> dict:
        parsed_dict = {}
        if "Sample" in self.parameters and "Notes" in self.parameters["Sample"]:
            notes = self.parameters["Sample"]["Notes"]
            pairs = notes.split(";")
            for pair in pairs:
                if ":" in pair:
                    key, val = pair.split(":", 1)
                    parsed_dict[key.strip()] = val.strip()
        return parsed_dict

    # Get input values from notes and map them to the experiment to be updated in excel sheet
    def map_input_values_from_notes_to_experiment(
        self, experiment: dict, initials: str
    ):
        if not self.notes or not experiment:
            return
        # Fill date from notes into excel sheet
        self.add_date_batch_id_to_excel(initials)

        # Get all keys experiment['inputs] and check if they are in notes
        # If they are, add them to experiment['inputs] with the value from notes
        for exp_key in experiment["inputs"]:
            exp_key_without_trailing_s = (
                exp_key[:-1] if exp_key.endswith("s") else exp_key
            )
            for notes_key in self.notes:
                # Check if Substring is present in the notes_key
                if exp_key_without_trailing_s in notes_key.lower():
                    # Get the [] inputs key from this experiment['inputs'][exp_key] dict
                    if experiment["inputs"][exp_key]:
                        for subKey in experiment["inputs"][exp_key]:
                            if "-inputs" in subKey.lower():
                                experiment["inputs"][exp_key][subKey] = self.notes[
                                    notes_key
                                ]
                                # Process metadata input further into subJson based on key
                                process_metadata_input(
                                    self.chemDB,
                                    experiment["inputs"][exp_key],
                                    subKey,
                                    exp_key,
                                    self.path,
                                )
                                break
                    break

    def add_date_batch_id_to_excel(self, initials: str):
        wb = load_workbook(filename=self.path, data_only=True)
        date = self.notes.get("Mix Date and time", None)
        polymerization_date = self.notes.get("Polymerization Date and time", None)
        sheet = wb["general"]
        if sheet and date:
            # Update B2 cell with date
            sheet["B2"] = date
            sheet["B3"] = polymerization_date
            # Update Initials
            sheet["B1"] = initials
            # Update batch ID
            parsed_date = datetime.strptime(str(date), "%m/%d/%Y %H:%M")
            # get the TYPE: IA/LDM from exp
            batch_id = f"{parsed_date.day}-{parsed_date.month}-{parsed_date.year} {parsed_date.strftime('%H:%M')}  {initials}"
            sheet["E32"] = batch_id
            # Save the workbook
            wb.save(self.path)


#     Fill the values
def process_metadata_input(
    chemDB: ChemDB,
    metadata_input: dict,
    subKey: str,
    excel_key: str,
    worksheet_path: str,
):

    if not chemDB:
        chemDB = ChemDB()

    records = metadata_input[subKey].split(", ")
    processed_records = []
    input_records = []

    for record in records:
        smile = ""
        abbrev, mass = record.split(" ")
        if abbrev and not chemDB.new_data.empty:
            full_name = chemDB.new_data.loc[
                chemDB.new_data["Abbreviation"] == abbrev, "Component"
            ].values[0]
            # new_data dataframe has index built on SMILES attribute
            smile = chemDB.new_data[chemDB.new_data["Abbreviation"] == abbrev].index[0]

        else:
            full_name = abbrev  # Default to the short name if not recognized

        processed_records.append({"name": full_name, "Measured mass (g)": float(mass)})
        if smile:
            logger.debug(f"SMILES: {smile} found in Chemistry database for {full_name}")
            input_records.append((full_name, smile, float(mass)))
        else:
            logger.error(f"SMILES not found in Chemistry database for {full_name}")
            input_records.append((full_name, abbrev, float(mass)))

    metadata_input[subKey] = processed_records
    # Update the template excel sheet with the values from notes
    update_excel_with_values(input_records, excel_key, worksheet_path)


# updates the excel sheet inplace for each input
def update_excel_with_values(values, input_sheet: str, path: str):
    wb = load_workbook(filename=path, data_only=True)
    if input_sheet == "" or not input_sheet:
        return
    sheet = wb[input_sheet]
    if not sheet:
        return

    # Update the sheet with new values
    for row, value in enumerate(values, start=2):  # Assuming the first row is headers
        full_name, abbrv, mass = value[0], value[1], value[2]
        sheet.cell(row=row, column=1, value=full_name)
        sheet.cell(row=row, column=2, value=abbrv)
        sheet.cell(row=row, column=3, value=float(mass))

    # Save the workbook
    wb.save(path)

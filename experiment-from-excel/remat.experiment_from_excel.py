#!/usr/bin/env python
import json

from chemistry import Monomer, ChemDB, Catalyst, Inhibitor, Solvent, Additive
import logging

import pyclowder.files
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pyclowder.extractors import Extractor
from pyclowder.utils import CheckMessage

moles_format = '{:.2e}'

def microliters_to_milli(value):
    if value and value != "-":
        return float(value)/1000.0
    else:
        return value

def mass_volume_conversion(value):
    if value == "-":
        return None 
    return value

def find_mass_column(row):
    if "Measured mass (g)" in row:
        return "Measured mass (g)"
    elif "Measured mass (mg)" in row:
        return "Measured mass (mg)"
    else:
        return None

def find_volume_column(row):
    if "Measured volume (μL)" in row:
        return "Measured volume (μL)"

def is_row_empty(row, input_title: str):

    mass_key = find_mass_column(row)
    volume_key = find_volume_column(row)

    smiles = row["SMILES"] if row["SMILES"] else None
    name = row["Name"] if row["Name"] else None
    mass = mass_volume_conversion(row[mass_key]) if mass_key else None 
    volume = mass_volume_conversion(row[volume_key]) if volume_key else None

    if name and smiles:
        if mass and volume:
            raise ValueError(f'On {input_title} tab: Only specify one of mass or volume in '+str(row))
        elif not mass and not volume:
            raise ValueError(f'On {input_title} tab: Volume or mass must be specified in '+str(row))
    
    if mass_key and volume_key:
        if not name and not smiles and not mass and not volume:
            return True
    elif mass_key:
        if not name and not smiles and not mass:
            return True
    elif volume_key:
        if not name and not smiles and not volume:
            return True

    if not all([row["Name"], row["SMILES"]]):
        raise ValueError(f'On {input_title} tab: Missing Name or SMILES in '+str(row))

    return False

def compute_values(inputs: dict):
    # First, create lists of input-specific chemistry converters with the observed values
    db = ChemDB()

    db.exists([compound["SMILES"] for compound in inputs['monomers']])
    monomers = {compound["SMILES"]:
        Monomer(compound["SMILES"], db, mass_volume_conversion(compound["Measured mass (g)"]),
                 microliters_to_milli(mass_volume_conversion(compound["Measured volume (μL)"]))) for compound in inputs['monomers']}

    print(monomers)

    db.exists([compound["SMILES"] for compound in inputs['catalysts']])
    catalysts = {compound["SMILES"]:
                     Catalyst(compound["SMILES"], db, mass_volume_conversion(compound["Measured mass (mg)"]) / 1000.0,
                              None) for compound in inputs['catalysts']}
    print(catalysts)

    db.exists([compound["SMILES"] for compound in inputs['inhibitors']])
    inhibitors = {compound["SMILES"]:
                      Inhibitor(compound["SMILES"], db, None,
                                microliters_to_milli(mass_volume_conversion(compound["Measured volume (μL)"]))) for compound in
                  inputs['inhibitors']
                  }
    print(inhibitors)

    db.exists([compound["SMILES"] for compound in inputs['additives']])
    additives = {compound["SMILES"]:
                     Additive(compound["SMILES"], db, mass_volume_conversion(compound["Measured mass (g)"]),
                              microliters_to_milli(mass_volume_conversion(compound["Measured volume (μL)"]))) for compound in
                 inputs['additives']
                 }
    print(additives)

    db.exists([compound["SMILES"] for compound in inputs['solvents']])
    solvents = {compound["SMILES"]:
                    Solvent(compound["SMILES"], db, mass_volume_conversion(compound["Measured mass (mg)"]),
                            microliters_to_milli(mass_volume_conversion(compound["Measured volume (μL)"]))) for compound in
                inputs['solvents']}
    print(solvents)

    # Now compute derived values (which requires knowledge of all of the inputs)
    monomer2 = [{
        "name": monomer["Name"],
        "SMILES": monomer["SMILES"],
        "Measured mass (g)": monomer["Measured mass (g)"],
        "Measured volume (μL)": monomer["Measured volume (μL)"],
        "Prepared in glovebox?": monomer["Prepared in glovebox?"],
        "Preparation temperature (°C)": monomer["Preparation temperature (°C)"],
        "Computed mass (g)": monomers[monomer["SMILES"]].mass,
        "Molecular Weight": monomers[monomer["SMILES"]].molecular_weight,
        "Moles": moles_format.format(monomers[monomer["SMILES"]].moles()),
        "Monomer mol%": monomers[monomer["SMILES"]].monomer_mol_percent(monomers.values())
    }
        for monomer in inputs["monomers"]]
    print("monomer2 --->", monomer2)

    catalyst2 = [{
        "name": catalyst["Name"],
        "SMILES": catalyst["SMILES"],
        "Measured mass (mg)": catalyst["Measured mass (mg)"],
        "Computed mass (g)": catalysts[catalyst["SMILES"]].mass,
        "Prepared in glovebox?": catalyst["Prepared in glovebox?"],
        "Preparation temperature (°C)": catalyst["Preparation temperature (°C)"],
        "Molecular Weight": catalysts[catalyst["SMILES"]].molecular_weight,
        "Moles": moles_format.format(catalysts[catalyst["SMILES"]].moles()),
        "Monomer:Catalyst molar ratio": catalysts[catalyst["SMILES"]].catalyst_monomer_molar_ratio(monomers.values())
    }
        for catalyst in inputs["catalysts"]]
    print("catalyst2 --->", catalyst2)

    inhibitor2 = [{
        "name": inhibitor["Name"],
        "SMILES": inhibitor["SMILES"],
        "Measured volume (μL)": inhibitor["Measured volume (μL)"],
        "Prepared in glovebox?": inhibitor["Prepared in glovebox?"],
        "Preparation temperature (°C)": inhibitor["Preparation temperature (°C)"],
        "Density": inhibitors[inhibitor["SMILES"]].density,
        "Computed mass (mg)": inhibitors[inhibitor["SMILES"]].mass,
        "Molecular Weight": inhibitors[inhibitor["SMILES"]].molecular_weight,
        "Moles": moles_format.format(inhibitors[inhibitor["SMILES"]].moles()),
        "Inhibitor:Catalyst molar ratio": inhibitors[inhibitor["SMILES"]].inhibitor_catalyst_molar_ratio(list(catalysts.values())[0])
    }
        for inhibitor in inputs["inhibitors"]]
    print("inhibitor2 --->", inhibitor2)

    additive2 = [{
        "name": additive["Name"],
        "SMILES": additive["SMILES"],
        "Measured mass (g)": additive["Measured mass (g)"],
        "Measured volume (μL)": additive["Measured volume (μL)"],
        "Computed mass (g)": additives[additive["SMILES"]].mass,
        "Molecular Weight": additives[additive["SMILES"]].molecular_weight,
        "Moles": moles_format.format(additives[additive["SMILES"]].moles()),
        "Wt Percent of Additives": additives[additive["SMILES"]].additive_weight_percent(
            list(additives.values()),
            list(monomers.values()),
            list(catalysts.values())[0],
            list(solvents.values())[0])
    }
        for additive in inputs["additives"]]
    print("additive2 --->", additive2)


    solvents2 = [{
        "name": solvent["Name"],
        "SMILES": solvent["SMILES"],
        "Measured mass (mg)": solvent["Measured mass (mg)"],
        "Measured volume (μL)": solvent["Measured volume (μL)"],
        "Computed mass (mg)": solvents[solvent["SMILES"]].mass,
        "Molecular Weight": solvents[solvent["SMILES"]].molecular_weight,
        "Moles": moles_format.format(solvents[solvent["SMILES"]].moles()),
        "Solvent concentration": solvents[solvent["SMILES"]].solvent_concentration(list(catalysts.values())[0])
    }
        for solvent in inputs["solvents"]]
    print("solvents2 --->", solvents2)

    inputs['monomers'] = monomer2
    inputs["catalysts"] = catalyst2
    inputs["inhibitors"] = inhibitor2
    inputs["additives"] = additive2
    inputs["solvents"] = solvents2

def read_inputs_from_worksheet(ws: Worksheet) -> list[dict]:
    inputs = []
    headers = [col.value for col in list(ws.rows)[0]]
    for row in ws.iter_rows(min_row=2):
        input_properties = {key: cell.value for key, cell in zip(headers, row)}

        if not is_row_empty(input_properties, input_title=ws.title):
            inputs.append(input_properties)

    return inputs

def read_procedure_from_worksheet(ws: Worksheet) -> dict:
    procedure = {}
    for row in ws.iter_rows():
        procedure[row[0].value] = row[1].value

    return procedure


def read_batch_id(wb: Workbook) -> str:
    batch_id_range = wb.defined_names["BatchID"].value.split("!")
    batch_id_cell = wb[batch_id_range[0]][batch_id_range[1]]
    return batch_id_cell.value



def excel_to_json(path):
    logger = logging.getLogger('__main__')

    wb = load_workbook(filename=path, data_only=True)
    # Find the data input spreadsheet version
    ss_version = [prop.value for prop
                  in wb.custom_doc_props.props
                  if prop.name == "File Version"][0]

    if ss_version != "1.0":
        logger.error(f"This extractor is not compatible with spreadsheet version {ss_version}")
        return {}

    inputs = {}
    batch_id = read_batch_id(wb)

    for sheet in wb.sheetnames:
        if sheet == "procedure":
            procedure = read_procedure_from_worksheet(wb[sheet])
        else:
            inputs[sheet] = read_inputs_from_worksheet(wb[sheet])

    compute_values(inputs)

    return {
        "Batch ID": batch_id,
        "procedure": procedure,
        "inputs": inputs
    }


class ExperimentFromExcel(Extractor):
    def __init__(self):
        Extractor.__init__(self)
        # parse command line and load default logging configuration
        self.setup()
        # setup logging for the exctractor
        logging.getLogger('pyclowder').setLevel(logging.DEBUG)
        logging.getLogger('__main__').setLevel(logging.DEBUG)

    def check_message(self, connector, host, secret_key, resource, parameters):
        logging.getLogger(__name__).debug("default check message : " + str(parameters))
        return CheckMessage.download

    def process_message(self, connector, host, secret_key, resource, parameters):
        logger = logging.getLogger('__main__')
        experiment = excel_to_json(resource['local_paths'][0])
        logger.debug(experiment)

        # store results as metadata
        metadata = {
            "@context": ["https://clowder.ncsa.illinois.edu/contexts/metadata.jsonld"],
            "dataset_id": resource['parent'].get('id', None),
            "content": experiment,
            "agent": {
                "@type": "cat:extractor",
                "extractor_id": host + "api/extractors/" + self.extractor_info['name']
            }
        }

        print(json.loads(json.dumps(metadata, default=str, ensure_ascii=False)))
        pyclowder.datasets.upload_metadata(connector, host, secret_key,
                                           resource['parent'].get('id', None), json.loads(json.dumps(metadata, default=str, ensure_ascii=False)))


if __name__ == "__main__":
    extractor = ExperimentFromExcel()
    extractor.start()
    # print(json.dumps(excel_to_json("/Users/bengal1/dev/MDF/clowder-extractors/experiment-from-excel/small_moles.xlsx")['inputs']['inhibitors'], default=str, ensure_ascii=False))